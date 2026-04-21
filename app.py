# app.py
# NOTA: Este es el servidor principal. Para iniciar: python app.py
# Las rutas controlan que ve cada usuario
# IMPORTANTE: El login es falso - cualquier usuario entra sin verificar contraseña

from flask import Flask, render_template, request, jsonify, redirect, url_for, session, send_file
from db import db, Usuario, Receta, Ingrediente, RecetaIngrediente, OrdenProduccion
from patrones import (
    ReporteFactory, InventarioObservable, 
    AlertasStock, LoggerSistema, ValidadorQuimico, 
    ValidacionBasica, ValidacionGramaje, ValidacionCompleta,
    crear_reporte_con_exportaciones
)
from datetime import datetime
from io import BytesIO
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///proquim.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.secret_key = "proquim_secret_key_2026"
# NOTA: Cambiar esta clave en produccion

# PATRON SINGLETON - Base de datos unica
db.init_app(app)

# PATRON OBSERVER - Sistema de alertas
observable = InventarioObservable()
observable.agregar(AlertasStock())
observable.agregar(LoggerSistema())

# PATRON STRATEGY - Validador quimico (se puede cambiar en tiempo real)
validador = ValidadorQuimico(ValidacionBasica())


# ============================================
# RUTA DE LOGIN (SIN VERIFICACION REAL)
# ============================================
# NOTA: Este login NO valida contraseña. Cualquier usuario entra.
# Solo para demostracion de patrones de diseno

@app.route('/')
def index():
    # Redirige al login por defecto
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        # IMPORTANTE: NO verifico nada - cualquier usuario entra
        username = request.form.get('username', 'demo')
        rol = request.form.get('rol', 'Administrador')
        
        # Guardo en sesion (opcional, para mostrar el nombre)
        session['user'] = username
        session['rol'] = rol
        
        return redirect(url_for('dashboard'))
    
    return render_template('login.html')

@app.route('/dashboard')
def dashboard():
    # NOTA: Si no hay sesion, igual entra (pero mostramos demo)
    user = session.get('user', 'Demo')
    rol = session.get('rol', 'Administrador')
    
    # PATRON OBSERVER - Verificar stock al cargar dashboard
    inventario = Ingrediente.query.all()
    for item in inventario:
        observable.verificar({'nombre': item.nombre, 'cantidad': item.cantidad, 'stock_minimo': item.stock_minimo})
    
    alertas = len([i for i in inventario if i.cantidad < i.stock_minimo])
    total_recetas = Receta.query.count()
    ordenes_activas = OrdenProduccion.query.filter_by(estado='Pendiente').count()
    
    return render_template('dashboard.html', 
                         user=user,
                         rol=rol,
                         total_recetas=total_recetas,
                         ordenes_activas=ordenes_activas,
                         alertas_stock=alertas)

@app.route('/recetas')
def recetas_page():
    # Muestra todas las recetas tecnicas
    recetas = Receta.query.all()
    # Convertir a dict para compatibilidad con templates
    recetas_dict = []
    for r in recetas:
        recetas_dict.append({
            'id': r.id,
            'nombre': r.nombre,
            'categoria': r.categoria,
            'rendimiento': r.rendimiento,
            'ingredientes': [{'nombre': ing.ingrediente.nombre, 'gramos': ing.gramos} for ing in r.ingredientes]
        })
    return render_template('recetas.html', recetas=recetas_dict)

@app.route('/validar_receta', methods=['POST'])
def validar_receta():
    # PATRON STRATEGY - Valida compatibilidad quimica
    data = request.get_json()
    ingredientes = data.get('ingredientes', [])
    resultado = validador.validar(ingredientes)
    return jsonify(resultado)

@app.route('/cambiar_validacion', methods=['POST'])
def cambiar_validacion():
    # NOTA: Permite cambiar la estrategia de validacion en tiempo real
    data = request.get_json()
    tipo = data.get('tipo', 'basica')
    if tipo == 'gramaje':
        validador.set_estrategia(ValidacionGramaje())
    elif tipo == 'completa':
        validador.set_estrategia(ValidacionCompleta())
    else:
        validador.set_estrategia(ValidacionBasica())
    return jsonify({"mensaje": f"Estrategia cambiada a: {tipo}"})

@app.route('/ordenes')
def ordenes_page():
    # Muestra las ordenes de produccion
    ordenes = OrdenProduccion.query.all()
    recetas = Receta.query.all()
    
    # Convertir a dict para compatibilidad
    ordenes_dict = []
    for o in ordenes:
        ordenes_dict.append({
            'id': o.id,
            'receta_id': o.receta_id,
            'receta_nombre': o.receta.nombre,
            'cantidad_producir': o.cantidad_producir,
            'estado': o.estado,
            'fecha': o.fecha.strftime('%Y-%m-%d'),
            'trazabilidad': {'lote_produccion': o.lote_produccion} if o.lote_produccion else None
        })
    
    recetas_dict = [{'id': r.id, 'nombre': r.nombre} for r in recetas]
    return render_template('ordenes_produccion.html', ordenes=ordenes_dict, recetas=recetas_dict)

@app.route('/crear_orden', methods=['POST'])
def crear_orden():
    # NOTA: Crea una nueva orden y la guarda en la BD
    data = request.get_json()
    receta_id = data.get('receta_id')
    cantidad = data.get('cantidad')
    
    receta = Receta.query.get(receta_id)
    if not receta:
        return jsonify({"error": "Receta no encontrada"}), 404
    
    nueva_orden = OrdenProduccion(
        receta_id=receta_id,
        cantidad_producir=cantidad,
        estado='Pendiente'
    )
    db.session.add(nueva_orden)
    db.session.commit()
    
    return jsonify({"mensaje": f"Orden creada: {receta.nombre} - {cantidad}kg", "orden": {"id": nueva_orden.id}})


@app.route('/inventario')
def inventario_page():
    # PATRON OBSERVER - Verificar stock al cargar inventario
    ingredientes = Ingrediente.query.all()
    for item in ingredientes:
        observable.verificar({'nombre': item.nombre, 'cantidad': item.cantidad, 'stock_minimo': item.stock_minimo})
    
    # Convertir a dict para compatibilidad
    inventario_dict = [{'id': i.id, 'nombre': i.nombre, 'cantidad': i.cantidad, 
                        'stock_minimo': i.stock_minimo, 'unidad': i.unidad} for i in ingredientes]
    return render_template('inventario.html', inventario=inventario_dict)

@app.route('/actualizar_stock', methods=['POST'])
def actualizar_stock():
    # NOTA: Actualiza cantidad de un material y dispara alertas si es necesario
    data = request.get_json()
    item_id = data.get('id')
    nueva_cantidad = data.get('cantidad')
    
    item = Ingrediente.query.get(item_id)
    if not item:
        return jsonify({"error": "Item no encontrado"}), 404
    
    item.cantidad = nueva_cantidad
    db.session.commit()
    observable.verificar({'nombre': item.nombre, 'cantidad': item.cantidad, 'stock_minimo': item.stock_minimo})
    
    return jsonify({"mensaje": f"Stock actualizado: {item.nombre} = {nueva_cantidad} {item.unidad}"})

@app.route('/reportes')
def reportes_page():
    # PATRON FACTORY METHOD + PATRON DECORATOR
    # Factory crea el reporte base, Decorator agrega exportaciones
    ordenes = OrdenProduccion.query.all()
    total = len(ordenes)
    completadas = len([o for o in ordenes if o.estado == 'Completada'])
    eficiencia = (completadas / total * 100) if total > 0 else 0
    
    factory = ReporteFactory()
    ordenes_dict = [{"estado": o.estado} for o in ordenes]
    reporte = factory.crear("produccion", {"ordenes": ordenes_dict})
    datos_reporte = reporte.generar()
    
    # PATRON DECORATOR - Agrega exportacion a PDF y Excel
    reporte_con_export = crear_reporte_con_exportaciones(datos_reporte, pdf=True, excel=True)
    reporte_final = reporte_con_export.generar()
    
    return render_template('reportes.html', 
                         total_ordenes=total,
                         ordenes_completadas=completadas,
                         eficiencia=round(eficiencia, 1),
                         reporte=reporte_final)

@app.route('/exportar_pdf')
def exportar_pdf():
    # NOTA: Genera un PDF real con los datos de órdenes de producción
    try:
        # Obtener datos de la BD
        ordenes = OrdenProduccion.query.all()
        total = len(ordenes)
        completadas = len([o for o in ordenes if o.estado == 'Completada'])
        eficiencia = (completadas / total * 100) if total > 0 else 0
        
        # Crear documento PDF en memoria
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        story = []
        styles = getSampleStyleSheet()
        
        # Estilo personalizado para título
        titulo_style = ParagraphStyle(
            'TituloPDF',
            parent=styles['Heading1'],
            fontSize=18,
            textColor='#1a3a52',
            spaceAfter=30,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        
        # Título
        story.append(Paragraph("REPORTE DE PRODUCCIÓN", titulo_style))
        story.append(Paragraph(f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal']))
        story.append(Spacer(1, 0.3*inch))
        
        # Resumen
        resumen_data = [
            ['Total de Órdenes', str(total)],
            ['Completadas', str(completadas)],
            ['Pendientes', str(total - completadas)],
            ['Eficiencia (%)', f"{round(eficiencia, 1)}%"]
        ]
        
        resumen_table = Table(resumen_data, colWidths=[3*inch, 2*inch])
        resumen_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), '#e8f4f8'),
            ('TEXTCOLOR', (0, 0), (-1, -1), '#000000'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 1, '#cccccc')
        ]))
        story.append(resumen_table)
        story.append(Spacer(1, 0.3*inch))
        
        # Tabla de órdenes detalladas
        story.append(Paragraph("Detalle de Órdenes", styles['Heading2']))
        story.append(Spacer(1, 0.2*inch))
        
        ordenes_data = [['ID', 'Receta', 'Cantidad', 'Estado', 'Fecha']]
        for orden in ordenes:
            ordenes_data.append([
                str(orden.id),
                orden.receta.nombre[:20],
                f"{orden.cantidad_producir} L/kg",
                orden.estado,
                orden.fecha.strftime('%d/%m/%Y')
            ])
        
        ordenes_table = Table(ordenes_data, colWidths=[0.6*inch, 2*inch, 1.2*inch, 1.2*inch, 1*inch])
        ordenes_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), '#1a3a52'),
            ('TEXTCOLOR', (0, 0), (-1, 0), 'white'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, '#cccccc'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), ['white', '#f0f0f0'])
        ]))
        story.append(ordenes_table)
        
        # Generar PDF
        doc.build(story)
        buffer.seek(0)
        
        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f'reporte_produccion_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        )
    
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/exportar_excel')
def exportar_excel():
    # NOTA: Genera un Excel real con los datos de órdenes y producción
    try:
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Reportes de Producción"
        
        # Estilos
        header_fill = PatternFill(start_color="1a3a52", end_color="1a3a52", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        title_font = Font(bold=True, size=14, color="1a3a52")
        
        # Título
        ws['A1'] = "REPORTE DE PRODUCCIÓN - PROQUIM"
        ws['A1'].font = title_font
        ws.merge_cells('A1:E1')
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Fecha de generación
        ws['A2'] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws['A2'].font = Font(italic=True)
        ws.merge_cells('A2:E2')
        
        ws.append([])  # Fila vacía
        
        # ===== HOJA 1: RESUMEN =====
        ws['A4'] = "RESUMEN"
        ws['A4'].font = Font(bold=True, size=12, color="1a3a52")
        
        ordenes = OrdenProduccion.query.all()
        total = len(ordenes)
        completadas = len([o for o in ordenes if o.estado == 'Completada'])
        eficiencia = (completadas / total * 100) if total > 0 else 0
        
        ws['A5'] = "Total de Órdenes"
        ws['B5'] = total
        ws['A6'] = "Órdenes Completadas"
        ws['B6'] = completadas
        ws['A7'] = "Órdenes Pendientes"
        ws['B7'] = total - completadas
        ws['A8'] = "Eficiencia (%)"
        ws['B8'] = round(eficiencia, 1)
        
        ws.append([])
        
        # ===== HOJA 1: DETALLE DE ÓRDENES =====
        ws['A10'] = "DETALLE DE ÓRDENES"
        ws['A10'].font = Font(bold=True, size=12, color="1a3a52")
        
        # Headers de tabla
        headers = ['ID', 'Receta', 'Cantidad', 'Estado', 'Fecha', 'Lote']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=11, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Datos de órdenes
        row = 12
        for orden in ordenes:
            ws.cell(row=row, column=1).value = orden.id
            ws.cell(row=row, column=2).value = orden.receta.nombre
            ws.cell(row=row, column=3).value = orden.cantidad_producir
            ws.cell(row=row, column=4).value = orden.estado
            ws.cell(row=row, column=5).value = orden.fecha.strftime('%d/%m/%Y')
            ws.cell(row=row, column=6).value = orden.lote_produccion or "N/A"
            row += 1
        
        # ===== HOJA 2: INVENTARIO =====
        ws2 = wb.create_sheet("Inventario")
        ws2['A1'] = "INVENTARIO - PROQUIM"
        ws2['A1'].font = title_font
        ws2.merge_cells('A1:D1')
        ws2['A1'].alignment = Alignment(horizontal='center')
        
        # Headers
        headers_inv = ['Ingrediente', 'Cantidad', 'Stock Mínimo', 'Unidad']
        for col, header in enumerate(headers_inv, 1):
            cell = ws2.cell(row=3, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Datos de inventario
        ingredientes = Ingrediente.query.all()
        row = 4
        for ing in ingredientes:
            ws2.cell(row=row, column=1).value = ing.nombre
            ws2.cell(row=row, column=2).value = ing.cantidad
            ws2.cell(row=row, column=3).value = ing.stock_minimo
            ws2.cell(row=row, column=4).value = ing.unidad
            
            # Colorear si está bajo stock
            if ing.cantidad < ing.stock_minimo:
                for col in range(1, 5):
                    ws2.cell(row=row, column=col).fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
            
            row += 1
        
        # ===== HOJA 3: RECETAS =====
        ws3 = wb.create_sheet("Recetas")
        ws3['A1'] = "RECETAS TÉCNICAS - PROQUIM"
        ws3['A1'].font = title_font
        ws3.merge_cells('A1:D1')
        ws3['A1'].alignment = Alignment(horizontal='center')
        
        recetas = Receta.query.all()
        row = 3
        for receta in recetas:
            ws3[f'A{row}'] = f"Receta: {receta.nombre}"
            ws3[f'A{row}'].font = Font(bold=True, color="1a3a52")
            ws3.merge_cells(f'A{row}:D{row}')
            
            row += 1
            ws3[f'A{row}'] = "Categoría"
            ws3[f'B{row}'] = receta.categoria
            
            row += 1
            ws3[f'A{row}'] = "Rendimiento"
            ws3[f'B{row}'] = f"{receta.rendimiento} L/kg"
            
            row += 1
            ws3[f'A{row}'] = "Ingredientes"
            ws3[f'A{row}'].font = Font(bold=True)
            
            row += 1
            for ing in receta.ingredientes:
                ws3[f'A{row}'] = f"- {ing.ingrediente.nombre}"
                ws3[f'B{row}'] = f"{ing.gramos} g"
                row += 1
            
            row += 1  # Espaciado
        
        # Ajustar anchos de columna
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        
        ws2.column_dimensions['A'].width = 25
        ws2.column_dimensions['B'].width = 15
        ws2.column_dimensions['C'].width = 15
        ws2.column_dimensions['D'].width = 12
        
        ws3.column_dimensions['A'].width = 30
        ws3.column_dimensions['B'].width = 20
        ws3.column_dimensions['C'].width = 15
        ws3.column_dimensions['D'].width = 15
        
        # Guardar en memoria
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return send_file(
            buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'reporte_proquim_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
    
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/configuracion')
def configuracion_page():
    # NOTA: Obtener usuarios de la BD
    usuarios = Usuario.query.all()
    usuarios_dict = [{'username': u.username, 'password': u.password, 'rol': u.rol} for u in usuarios]
    return render_template('configuracion.html', usuarios=usuarios_dict, rol=session.get('rol', 'Administrador'))

@app.route('/crear_usuario', methods=['POST'])
def crear_usuario():
    # NOTA: Agrega un nuevo usuario a la BD
    data = request.get_json()
    nuevo_usuario = Usuario(
        username=data.get('username'),
        password=data.get('password'),
        rol=data.get('rol')
    )
    db.session.add(nuevo_usuario)
    db.session.commit()
    return jsonify({"mensaje": f"Usuario creado: {nuevo_usuario.username}"})

@app.route('/logout')
def logout():
    # NOTA: Cierra la sesion y borra los datos del usuario
    session.clear()
    return redirect(url_for('login'))


if __name__ == '__main__':
    # NOTA: debug=True solo para desarrollo. En produccion poner debug=False
    with app.app_context():
        # Crear tablas si no existen
        db.create_all()
    
    print("\n" + "="*60)
    print("PROQUIM - Sistema de Gestion de Produccion")
    print("="*60)
    print("✅ Base de datos: SQLite (proquim.db)")
    print("Singleton - Base de datos unica")
    print("Factory Method - Creacion de reportes")
    print("Observer - Alertas de inventario")
    print("Strategy - Validacion quimica")
    print("Decorator - Exportacion de reportes")
    print("="*60)
    print("IMPORTANTE: LOGIN FALSO - CUALQUIER USUARIO ENTRA")
    print("Servidor: http://127.0.0.1:5000")
    print("="*60 + "\n")
    app.run(debug=True)